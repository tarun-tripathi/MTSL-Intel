"""
Motherson Investment Intelligence
Streamlit application — 5 pages
streamlit run app.py
FIXED: Plotly duplicate key errors, chatbot chart loop keys
"""

import logging
import sys
from pathlib import Path

import io
import pandas as pd
import plotly.express as px
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT / "src"))
load_dotenv(ROOT / ".env")

st.set_page_config(
    page_title="Motherson Investment Intelligence",
    layout="wide",
    initial_sidebar_state="expanded",
)

from database import get_engine, get_db_stats
from chatbot import InvestmentChatbot, format_large_number

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("app")

CHART_HEIGHT = 360

def is_cloud_env() -> bool:
    """True if running on Streamlit Cloud with Gemini API key configured."""
    try:
        return "GEMINI_API_KEY" in st.secrets
    except Exception:
        return False

# ── Data loaders ─────────────────────────────────────────────

@st.cache_resource
def get_db_engine():
    return get_engine()

@st.cache_data(ttl=300)
def fetch_investments(_engine):
    try:
        return pd.read_sql("SELECT * FROM investments", _engine)
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def fetch_budget(_engine):
    try:
        return pd.read_sql("SELECT * FROM investment_budget", _engine)
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def fetch_cashflow(_engine):
    try:
        return pd.read_sql(
            "SELECT * FROM investment_monthly_cashflow ORDER BY period_date", _engine
        )
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_validation_report():
    path = ROOT / "data/logs/validation_report.csv"
    return pd.read_csv(path) if path.exists() else pd.DataFrame()

def metric_card(label, value, delta=None):
    st.markdown(
        f"""
        <div style="background:#f9fafb;padding:14px 18px;border-radius:10px;
                    margin-bottom:8px;border-left:3px solid #dc2626;">
            <div style="font-size:12px;color:#6b7280;margin-bottom:4px">{label}</div>
            <div style="font-size:22px;font-weight:600;color:#111827">{value}</div>
            {f'<div style="font-size:12px;color:#9ca3af;margin-top:2px">{delta}</div>' if delta else ''}
        </div>
        """,
        unsafe_allow_html=True,
    )


# ── Currency formatting ───────────────────────────────────────

def format_currency(value_k, currency_fmt: str = "us") -> str:
    """
    Format a k-EUR value into a human-readable string.
    currency_fmt='us'  → € 1,234.56M  (comma thousands, dot decimal)
    currency_fmt='eu'  → € 1.234,56M  (dot thousands, comma decimal)
    """
    try:
        v = float(value_k) * 1000
    except Exception:
        return str(value_k)

    if abs(v) >= 1_000_000_000:
        s = f"€ {v / 1_000_000_000:.2f}B"
    elif abs(v) >= 1_000_000:
        s = f"€ {v / 1_000_000:.2f}M"
    elif abs(v) >= 1_000:
        n = v / 1_000
        s = f"€ {n:.0f}K" if n == int(n) else f"€ {n:.1f}K"
    else:
        s = f"€ {v:,.2f}"

    if currency_fmt == "eu":
        # swap separators: 1,234.56 → 1.234,56
        s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return s


# ── Excel export ───────────────────────────────────────────────

def to_excel_bytes(df: pd.DataFrame, sheet_title: str = "Results") -> bytes:
    """Return a formatted .xlsx file as bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_title[:31])
        ws = writer.sheets[sheet_title[:31]]
        red_fill = PatternFill("solid", fgColor="DC2626")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
    return output.getvalue()


# ── PDF export ─────────────────────────────────────────────────

def to_pdf_bytes(df: pd.DataFrame, title: str, answer: str = "") -> bytes:
    """Return a simple PDF report as bytes using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    elements = []

    # Title + subtitle
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(
        "Motherson Investment Intelligence — generated report",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 8))

    # Answer summary (strip markdown bold markers)
    if answer:
        clean = answer.replace("**", "").replace("_", "")
        elements.append(Paragraph(clean, styles["Normal"]))
        elements.append(Spacer(1, 10))

    # Table — truncate to 200 rows so PDF stays readable
    display_df = df.head(200)
    header = list(display_df.columns)
    rows   = [list(map(str, r)) for r in display_df.values.tolist()]
    data   = [header] + rows

    col_count = len(header)
    page_w    = landscape(A4)[0] - 72          # usable width
    col_w     = page_w / max(col_count, 1)

    t = Table(data, colWidths=[col_w] * col_count, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#DC2626")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    if len(df) > 200:
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(
            f"Showing first 200 of {len(df):,} rows.",
            styles["Normal"],
        ))

    doc.build(elements)
    return output.getvalue()


# ── Sidebar ───────────────────────────────────────────────────

with st.sidebar:
    st.markdown(
        """
        <div style="padding:8px 0 16px 0">
            <span style="font-size:22px;font-weight:700;color:#dc2626;
                         letter-spacing:2px">MTSL</span>
            <span style="font-size:13px;color:#6b7280;margin-left:8px">Intel</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption("Investment Planning Intelligence")
    st.divider()

    page = st.radio(
        "Navigate",
        ["Dashboard", "Data Explorer", "Validation Report", "Chatbot", "Pipeline"],
        label_visibility="collapsed",
    )

    st.divider()

    try:
        engine = get_db_engine()
        stats  = get_db_stats(engine)
        st.markdown("**Database**")
        st.caption("Connected — PostgreSQL")
        st.caption(f"{stats.get('investments', 0):,} records")
    except RuntimeError as e:
        st.markdown("**Database**")
        st.error(str(e))

    st.divider()

    st.markdown("**LLM**")
    if is_cloud_env():
        st.caption("Google Gemini 1.5 Flash")
    else:
        try:
            import requests
            requests.get("http://localhost:11434", timeout=2)
            st.caption("Ollama running — Mistral")
        except Exception:
            st.caption("Ollama not running")
            st.caption("Start with: `ollama serve`")

    st.divider()
    st.markdown("**Display**")
    _fmt_choice = st.radio(
        "Currency format",
        ["US  (1,000.00)", "EU  (1.000,00)"],
        horizontal=True,
        label_visibility="collapsed",
    )
    st.session_state.currency_fmt = "eu" if "EU" in _fmt_choice else "us"

    st.divider()
    st.caption("v1.1  |  Motherson Investment Intelligence")


# ── Dashboard ─────────────────────────────────────────────────

if page == "Dashboard":
    st.title("Investment Planning Dashboard")
    st.caption("Motherson Group — 5-year investment plan 2025/26 to 2029/30")

    try:
        engine = get_db_engine()
    except Exception:
        st.error("Database not connected. Run the pipeline first.")
        st.stop()

    df_inv  = fetch_investments(engine)
    df_bud  = fetch_budget(engine)
    df_cash = fetch_cashflow(engine)

    if df_inv.empty:
        st.warning("No data available. Run the pipeline.")
        st.stop()

    df = df_inv.merge(df_bud, on="row_id", how="left")

    # KPIs
    st.subheader("Key metrics")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        metric_card("Total investments", f"{len(df_inv):,}")
    with c2:
        total = df_bud["total_5y_k_eur"].sum() if "total_5y_k_eur" in df_bud.columns else 0
        cfmt  = st.session_state.get("currency_fmt", "us")
        metric_card("5-year total", format_currency(total, cfmt), "all fiscal years")
    with c3:
        fy   = df_bud["budget_fy_2526"].sum() if "budget_fy_2526" in df_bud.columns else 0
        cfmt = st.session_state.get("currency_fmt", "us")
        metric_card("FY 2025/26", format_currency(fy, cfmt), "current year")
    with c4:
        metric_card("Plants", str(df_inv["plant"].nunique()), "global sites")
    with c5:
        metric_card("Regions", str(df_inv["region"].nunique()), "geographic")

    st.divider()

    # Row 1
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Budget by company — FY 2025/26")
        if "budget_fy_2526" in df.columns:
            co = df.groupby("company")["budget_fy_2526"].sum().reset_index()
            co.columns = ["Company", "Budget (k€)"]
            co = co.sort_values("Budget (k€)", ascending=False)
            fig = px.bar(
                co, x="Company", y="Budget (k€)",
                text=co["Budget (k€)"].apply(format_large_number),
                color_discrete_sequence=["#dc2626"],
            )
            fig.update_layout(
                showlegend=False, plot_bgcolor="white",
                height=CHART_HEIGHT, font_family="sans-serif",
            )
            fig.update_traces(textposition="outside")
            fig.update_xaxes(title=None)
            # FIX: unique key
            st.plotly_chart(fig, use_container_width=True, key="dashboard_company")

    with col2:
        st.subheader("Investment by category")
        cat = df_inv["investment_category"].value_counts().reset_index()
        cat.columns = ["Category", "Count"]
        fig = px.pie(
            cat, names="Category", values="Count",
            color_discrete_sequence=px.colors.qualitative.Set2,
            hole=0.4,
        )
        fig.update_layout(height=CHART_HEIGHT)
        # FIX: unique key
        st.plotly_chart(fig, use_container_width=True, key="dashboard_category")

    # Row 2
    col3, col4 = st.columns(2)
    with col3:
        st.subheader("Budget by region — FY 2025/26")
        if "budget_fy_2526" in df.columns:
            reg = df.groupby("region")["budget_fy_2526"].sum().reset_index()
            reg.columns = ["Region", "Budget (k€)"]
            reg = reg.sort_values("Budget (k€)", ascending=True)
            fig = px.bar(
                reg, x="Budget (k€)", y="Region", orientation="h",
                text=reg["Budget (k€)"].apply(format_large_number),
                color_discrete_sequence=["#1d4ed8"],
            )
            fig.update_layout(
                showlegend=False, plot_bgcolor="white",
                height=CHART_HEIGHT,
            )
            fig.update_traces(textposition="outside")
            fig.update_yaxes(title=None)
            # FIX: unique key
            st.plotly_chart(fig, use_container_width=True, key="dashboard_region")

    with col4:
        st.subheader("5-year budget trend")
        fy_cols = {
            "FY 2025/26": "budget_fy_2526",
            "FY 2026/27": "budget_fy_2627",
            "FY 2027/28": "plan_fy_2728",
            "FY 2028/29": "plan_fy_2829",
            "FY 2029/30": "plan_fy_2930",
        }
        trend = [
            {"Year": k, "Budget (k€)": df_bud[v].sum()}
            for k, v in fy_cols.items() if v in df_bud.columns
        ]
        if trend:
            tdf = pd.DataFrame(trend)
            fig = px.line(
                tdf, x="Year", y="Budget (k€)", markers=True,
                color_discrete_sequence=["#dc2626"],
                text=tdf["Budget (k€)"].apply(format_large_number),
            )
            fig.update_traces(textposition="top center")
            fig.update_layout(plot_bgcolor="white", height=CHART_HEIGHT)
            fig.update_xaxes(title=None)
            # FIX: unique key
            st.plotly_chart(fig, use_container_width=True, key="dashboard_trend")

    # Row 3
    col5, col6 = st.columns(2)
    with col5:
        st.subheader("Source of funding")
        fund = df_inv["source_of_funding"].value_counts().reset_index()
        fund.columns = ["Source", "Count"]
        fig = px.pie(
            fund, names="Source", values="Count",
            color_discrete_sequence=px.colors.qualitative.Pastel,
            hole=0.3,
        )
        fig.update_layout(height=CHART_HEIGHT)
        # FIX: unique key
        st.plotly_chart(fig, use_container_width=True, key="dashboard_funding")

    with col6:
        st.subheader("Top 10 plants by investment count")
        top = df_inv["plant"].value_counts().head(10).reset_index()
        top.columns = ["Plant", "Count"]
        fig = px.bar(
            top, x="Count", y="Plant", orientation="h",
            color_discrete_sequence=["#0891b2"],
            text="Count",
        )
        fig.update_layout(
            plot_bgcolor="white", height=CHART_HEIGHT,
            yaxis={"categoryorder": "total ascending"},
        )
        fig.update_yaxes(title=None)
        fig.update_traces(textposition="outside")
        # FIX: unique key
        st.plotly_chart(fig, use_container_width=True, key="dashboard_plants")


# ── Data Explorer ─────────────────────────────────────────────

elif page == "Data Explorer":
    st.title("Data Explorer")
    st.caption("Browse and filter investment records")

    try:
        engine = get_db_engine()
    except Exception:
        st.error("Database not connected.")
        st.stop()

    df_inv = fetch_investments(engine)
    df_bud = fetch_budget(engine)

    if df_inv.empty:
        st.warning("No data available.")
        st.stop()

    df = df_inv.merge(
        df_bud[["row_id", "budget_fy_2526", "budget_fy_2627", "total_5y_k_eur"]],
        on="row_id", how="left",
    )

    with st.expander("Filters", expanded=True):
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            companies = ["All"] + sorted(df["company"].dropna().unique().tolist())
            sel_co = st.selectbox("Company", companies)
        with fc2:
            regions = ["All"] + sorted(df["region"].dropna().unique().tolist())
            sel_reg = st.selectbox("Region", regions)
        with fc3:
            cats = ["All"] + sorted(df["investment_category"].dropna().unique().tolist())
            sel_cat = st.selectbox("Category", cats)
        with fc4:
            plants = ["All"] + sorted(df["plant"].dropna().unique().tolist())
            sel_plant = st.selectbox("Plant", plants)

    filtered = df.copy()
    if sel_co    != "All": filtered = filtered[filtered["company"] == sel_co]
    if sel_reg   != "All": filtered = filtered[filtered["region"]  == sel_reg]
    if sel_cat   != "All": filtered = filtered[filtered["investment_category"] == sel_cat]
    if sel_plant != "All": filtered = filtered[filtered["plant"]   == sel_plant]

    st.caption(f"{len(filtered):,} records")

    display_cols = [c for c in [
        "investment_id", "company", "region", "plant",
        "investment_category", "customer", "investment_description",
        "source_of_funding", "tangible_intangible",
        "budget_fy_2526", "budget_fy_2627", "total_5y_k_eur",
    ] if c in filtered.columns]

    st.dataframe(filtered[display_cols].reset_index(drop=True), use_container_width=True, height=500)

    csv = filtered[display_cols].to_csv(index=False).encode("utf-8")
    st.download_button("Download as CSV", csv, "investments.csv", "text/csv")


# ── Validation Report ─────────────────────────────────────────

elif page == "Validation Report":
    st.title("Validation Report")
    st.caption("Data quality check — 12 business rules")

    report_df = load_validation_report()

    if report_df.empty:
        st.warning("Validation report not found. Run the pipeline first.")
        st.stop()

    passed = int((report_df["failed"] == 0).sum()) if "failed" in report_df.columns else 0
    failed = int((report_df["failed"]  > 0).sum()) if "failed" in report_df.columns else 0
    avg_pr = float(report_df["pass_rate"].mean()) if "pass_rate" in report_df.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1: metric_card("Rules run",    str(len(report_df)))
    with c2: metric_card("Rules passed", str(passed))
    with c3: metric_card("Rules failed", str(failed))
    with c4: metric_card("Pass rate",    f"{avg_pr:.1f}%")

    st.divider()

    cols = [c for c in ["status", "rule_id", "rule_name", "severity",
                         "passed", "failed", "pass_rate"] if c in report_df.columns]
    st.dataframe(report_df[cols].reset_index(drop=True), use_container_width=True)

    st.subheader("Pass rate per rule")
    fig = px.bar(
        report_df.sort_values("pass_rate"),
        x="pass_rate", y="rule_name", orientation="h",
        color="severity",
        color_discrete_map={"HARD": "#dc2626", "SOFT": "#f59e0b"},
        text="pass_rate",
    )
    fig.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
    fig.update_layout(plot_bgcolor="white", height=420, xaxis_range=[0, 105])
    fig.update_yaxes(title=None)
    # FIX: unique key
    st.plotly_chart(fig, use_container_width=True, key="validation_chart")

    quar_path = ROOT / "data/logs/quarantine.csv"
    if quar_path.exists():
        quar_df = pd.read_csv(quar_path)
        if len(quar_df):
            st.subheader(f"Quarantine — {len(quar_df)} rows")
            st.dataframe(quar_df.head(50), use_container_width=True)
        else:
            st.success("No rows quarantined. All records passed hard validation rules.")
    else:
        st.success("No rows quarantined.")


# ── Chatbot ───────────────────────────────────────────────────

elif page == "Chatbot":
    st.title("Investment Data Chatbot")
    st.caption("Ask questions about investment data in plain English, Hindi, Hinglish, or German.")
 
    if "chatbot" not in st.session_state:
        try:
            engine = get_db_engine()
            st.session_state.chatbot = InvestmentChatbot(engine=engine)
        except Exception as e:
            st.error(f"Could not initialise chatbot: {e}")
            st.stop()
 
    if "messages" not in st.session_state:
        st.session_state.messages = []
 
    chatbot = st.session_state.chatbot
 
    # Show warning only on local when Ollama is down. On cloud we use Gemini.
    if not is_cloud_env() and not chatbot.ollama_available:
        st.warning("Ollama is not running. Start it with: `ollama serve`")
 
    # ── Process suggestion click ──────────────────────────
    if "pending" in st.session_state:
        question = st.session_state.pop("pending")
        st.session_state.messages.append({"role": "user", "content": question})
        with st.spinner("Working on it..."):
            result = chatbot.ask(question)
        answer = result["answer"] or result.get("error", "Error")
        st.session_state.messages.append({
            "role":          "assistant",
            "content":       answer,
            "sql":           result["sql"],
            "chart":         result["chart_data"],
            "df":            result.get("df"),
            "csv_data":      result.get("csv_data"),
            "rows":          result.get("rows", 0),
            "clarification": result.get("clarification", False),
        })
        st.rerun()
 
    # ── Helper: render one chart ──────────────────────────
    def render_chart(chart, key):
        if not chart:
            return
        ct        = chart["chart_type"]
        val_col   = chart.get("value_col", "")
        is_budget = any(x in val_col.lower() for x in
                        ["budget","eur","amount","total","avg_budget"])

        # Convert k€ → M€ for budget columns so axis reads "€40M" not "40000"
        raw_vals = chart["values"]
        if is_budget and raw_vals and max((v or 0) for v in raw_vals) > 500:
            plot_vals  = [round((v or 0) / 1000, 2) for v in raw_vals]
            y_label    = f"{val_col} (M€)"
            y_tickpfx  = "€"
            y_ticksfx  = "M"
        else:
            plot_vals  = raw_vals
            y_label    = val_col
            y_tickpfx  = ""
            y_ticksfx  = ""

        if ct == "bar":
            fig = px.bar(
                x=chart["labels"], y=plot_vals,
                labels={"x": chart["label_col"], "y": y_label},
                color_discrete_sequence=["#dc2626"],
            )
        elif ct == "line":
            fig = px.line(
                x=chart["labels"], y=plot_vals,
                labels={"x": chart["label_col"], "y": y_label},
                markers=True,
                color_discrete_sequence=["#dc2626"],
            )
        else:
            fig = px.pie(names=chart["labels"], values=raw_vals)
        fig.update_layout(plot_bgcolor="white", height=320)
        if ct in ("bar", "line") and is_budget:
            fig.update_yaxes(tickprefix=y_tickpfx, ticksuffix=y_ticksfx,
                             tickformat=".1f")
        st.plotly_chart(fig, use_container_width=True, key=key)
 
    # ── Chat history ──────────────────────────────────────
    for i, msg in enumerate(st.session_state.messages):
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])
 
            if msg["role"] == "assistant":
                # Formatted table
                df_msg = msg.get("df")
                if df_msg is not None and not df_msg.empty:
                    st.dataframe(df_msg, use_container_width=True, hide_index=True)
                    rows = msg.get("rows", len(df_msg))
                    st.caption(f"{rows:,} records")
 
                # Chart
                if msg.get("chart") and not msg.get("clarification"):
                    render_chart(msg["chart"], key=f"chat_chart_{i}")
 
                # Download buttons — CSV, Excel, PDF
                raw_df = msg.get("df")
                csv    = msg.get("csv_data")
                if csv and raw_df is not None and not raw_df.empty:
                    dl_cols = st.columns(3)
                    with dl_cols[0]:
                        st.download_button(
                            label="⬇ CSV",
                            data=csv,
                            file_name="motherson_data.csv",
                            mime="text/csv",
                            key=f"dl_csv_{i}",
                            use_container_width=True,
                        )
                    with dl_cols[1]:
                        st.download_button(
                            label="⬇ Excel",
                            data=to_excel_bytes(raw_df),
                            file_name="motherson_data.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_xlsx_{i}",
                            use_container_width=True,
                        )
                    with dl_cols[2]:
                        st.download_button(
                            label="⬇ PDF",
                            data=to_pdf_bytes(raw_df, "Motherson Investment Data", msg.get("content", "")),
                            file_name="motherson_data.pdf",
                            mime="application/pdf",
                            key=f"dl_pdf_{i}",
                            use_container_width=True,
                        )
 
                # SQL expander
                if msg.get("sql"):
                    with st.expander("SQL query"):
                        st.code(msg["sql"], language="sql")
 
    # ── New input ─────────────────────────────────────────
    if question := st.chat_input("Ask about the investment data..."):
        st.session_state.messages.append({"role": "user", "content": question})
 
        with st.chat_message("user"):
            st.markdown(question)
 
        with st.chat_message("assistant"):
            with st.spinner("Working on it..."):
                result = chatbot.ask(question)
 
            answer = result["answer"] or result.get("error", "Error")
            st.markdown(answer)
 
            df_res = result.get("df")
            if df_res is not None and not df_res.empty:
                st.dataframe(df_res, use_container_width=True, hide_index=True)
                st.caption(f"{result['rows']:,} records")
 
            if result.get("chart_data") and not result.get("clarification"):
                render_chart(
                    result["chart_data"],
                    key=f"chat_chart_new_{len(st.session_state.messages)}"
                )
 
            if result.get("csv_data") and df_res is not None and not df_res.empty:
                _k = len(st.session_state.messages)
                dl_cols = st.columns(3)
                with dl_cols[0]:
                    st.download_button(
                        label="⬇ CSV",
                        data=result["csv_data"],
                        file_name="motherson_data.csv",
                        mime="text/csv",
                        key=f"dl_new_csv_{_k}",
                        use_container_width=True,
                    )
                with dl_cols[1]:
                    st.download_button(
                        label="⬇ Excel",
                        data=to_excel_bytes(df_res),
                        file_name="motherson_data.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_new_xlsx_{_k}",
                        use_container_width=True,
                    )
                with dl_cols[2]:
                    st.download_button(
                        label="⬇ PDF",
                        data=to_pdf_bytes(df_res, "Motherson Investment Data", answer),
                        file_name="motherson_data.pdf",
                        mime="application/pdf",
                        key=f"dl_new_pdf_{_k}",
                        use_container_width=True,
                    )
 
            if result.get("sql"):
                with st.expander("SQL query"):
                    st.code(result["sql"], language="sql")
 
        st.session_state.messages.append({
            "role":          "assistant",
            "content":       answer,
            "sql":           result["sql"],
            "chart":         result["chart_data"],
            "df":            result.get("df"),
            "csv_data":      result.get("csv_data"),
            "rows":          result.get("rows", 0),
            "clarification": result.get("clarification", False),
        })
 
    if st.session_state.messages:
        if st.button("Clear conversation"):
            st.session_state.messages = []
            if "chatbot" in st.session_state:
                st.session_state.chatbot.clear_memory()
            st.rerun()
 


# ── Pipeline ──────────────────────────────────────────────────

elif page == "Pipeline":
    st.title("Data Pipeline")
    st.caption("Run the full ingestion, cleaning, validation and database load.")

    excel_path = ROOT / "data/raw/Tarun_-_Intern_Assignment-Data.xlsx"

    if excel_path.exists():
        st.success(f"Source file found — {excel_path.name} ({excel_path.stat().st_size/1024/1024:.1f} MB)")
    else:
        st.error("Excel file not found in data/raw/")

    try:
        engine = get_db_engine()
        stats  = get_db_stats(engine)
        st.info(f"Database connected — {stats.get('investments', 0):,} investments currently loaded")
    except Exception as e:
        st.warning(f"Database: {e}")

    st.divider()

    reload = st.checkbox("Drop and recreate tables on run", value=True)

    if st.button("Run pipeline", type="primary"):
        if not excel_path.exists():
            st.error("Excel file not found.")
        else:
            with st.spinner("Running pipeline — this may take 30 to 60 seconds..."):
                try:
                    from pipeline import run_pipeline
                    summary = run_pipeline(str(excel_path), reload=reload)
                    st.success("Pipeline completed successfully.")

                    val = summary.get("validation", {})
                    ing = summary.get("ingestion", {})
                    c1, c2, c3, c4 = st.columns(4)
                    with c1: st.metric("Rows ingested", f"{ing.get('rows', 0):,}")
                    with c2: st.metric("Valid rows",    f"{val.get('valid_rows', 0):,}")
                    with c3: st.metric("Quarantined",   f"{val.get('quarantine_rows', 0):,}")
                    with c4: st.metric("Pass rate",     f"{val.get('overall_pass_rate', 0):.1f}%")

                    fetch_investments.clear()
                    fetch_budget.clear()
                    fetch_cashflow.clear()

                except Exception as e:
                    st.error(f"Pipeline error: {e}")
                    import traceback
                    st.code(traceback.format_exc())

    st.divider()
    st.subheader("Log files")
    for fname, label in [
        ("data/logs/pipeline.log",         "Pipeline log"),
        ("data/logs/cleaning_log.csv",      "Cleaning audit trail"),
        ("data/logs/validation_report.csv", "Validation report"),
        ("data/logs/quarantine.csv",        "Quarantine data"),
    ]:
        p = ROOT / fname
        if p.exists():
            with open(p) as f:
                st.download_button(label, f.read(), p.name, key=f"dl_{p.name}")
        else:
            st.caption(f"Not yet generated — {label}")
